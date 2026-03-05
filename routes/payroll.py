from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from datetime import date, datetime
from extentions import db
from models import Employee, Attendance, Advance, Payroll
from sqlalchemy import func
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

payroll_bp = Blueprint('payroll', __name__)

MONTH_NAMES = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']


def calculate_payroll(year, month):
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    results = []

    for emp in employees:
        att = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            func.extract('year', Attendance.date) == year,
            func.extract('month', Attendance.date) == month
        ).all()

        present_days = sum(1 for a in att if a.status == 'present')
        half_days = sum(1 for a in att if a.status == 'half_day')
        effective_days = present_days + half_days * 0.5

        total_advance = db.session.query(func.sum(Advance.amount)).filter(
            Advance.employee_id == emp.id,
            func.extract('year', Advance.date) == year,
            func.extract('month', Advance.date) == month
        ).scalar() or 0

        gross = float(emp.daily_wage) * effective_days
        net = gross - float(total_advance)

        paid_record = Payroll.query.filter_by(
            employee_id=emp.id, month=month, year=year
        ).first()

        # Partial payment support
        is_partial = False
        partial_days = 0
        partial_amount = 0
        pending_days = 0
        pending_amount = 0

        if paid_record and not paid_record.is_paid and paid_record.paid_at is not None:
            # A partial payment has been recorded (paid_at set but is_paid=False)
            partial_days = float(paid_record.present_days)
            partial_amount = float(paid_record.net_salary)
            pending_days = round(effective_days - partial_days, 1)
            pending_amount = round(net - partial_amount, 2)
            is_partial = True

        results.append({
            'employee_id': emp.id,
            'name': emp.name,
            'daily_wage': float(emp.daily_wage),
            'present_days': present_days,
            'half_days': half_days,
            'effective_days': effective_days,
            'gross_salary': gross,
            'advance_deduction': float(total_advance),
            'net_salary': net,
            'is_paid': paid_record.is_paid if paid_record else False,
            'is_partial': is_partial,
            'partial_days': partial_days,
            'partial_amount': partial_amount,
            'pending_days': pending_days,
            'pending_amount': pending_amount,
        })

    return results


@payroll_bp.route('/payroll')
@login_required
def index():
    today = date.today()
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))

    payroll_data = calculate_payroll(year, month)

    total_gross = sum(p['gross_salary'] for p in payroll_data)
    total_advance = sum(p['advance_deduction'] for p in payroll_data)
    total_net = sum(p['net_salary'] for p in payroll_data)

    months = [{'value': i, 'label': MONTH_NAMES[i-1]} for i in range(1, 13)]

    return render_template('pages/payroll.html',
        payroll=payroll_data,
        total_gross=total_gross,
        total_advance=total_advance,
        total_net=total_net,
        selected_month=month,
        selected_year=year,
        month_name=MONTH_NAMES[month-1],
        months=months,
        current_year=today.year
    )


@payroll_bp.route('/payroll/mark-paid', methods=['POST'])
@login_required
def mark_paid():
    employee_id = int(request.form.get('employee_id'))
    month = int(request.form.get('month'))
    year = int(request.form.get('year'))

    data = calculate_payroll(year, month)
    emp_data = next((p for p in data if p['employee_id'] == employee_id), None)
    if not emp_data:
        flash('Employee not found.', 'error')
        return redirect(url_for('payroll.index', month=month, year=year))

    existing = Payroll.query.filter_by(employee_id=employee_id, month=month, year=year).first()
    if existing:
        existing.is_paid = True
        existing.present_days = emp_data['effective_days']
        existing.gross_salary = emp_data['gross_salary']
        existing.advance_deduction = emp_data['advance_deduction']
        existing.net_salary = emp_data['net_salary']
        existing.paid_at = datetime.now()
    else:
        record = Payroll(
            employee_id=employee_id, month=month, year=year,
            present_days=emp_data['effective_days'],
            daily_wage=emp_data['daily_wage'],
            gross_salary=emp_data['gross_salary'],
            advance_deduction=emp_data['advance_deduction'],
            net_salary=emp_data['net_salary'],
            is_paid=True, paid_at=datetime.now()
        )
        db.session.add(record)

    db.session.commit()
    flash(f'{emp_data["name"]} marked as fully paid!', 'success')
    return redirect(url_for('payroll.index', month=month, year=year))


@payroll_bp.route('/payroll/partial-pay', methods=['POST'])
@login_required
def partial_pay():
    employee_id = int(request.form.get('employee_id'))
    month = int(request.form.get('month'))
    year = int(request.form.get('year'))
    paid_days = float(request.form.get('paid_days', 0))

    data = calculate_payroll(year, month)
    emp_data = next((p for p in data if p['employee_id'] == employee_id), None)

    if not emp_data:
        flash('Employee not found.', 'error')
        return redirect(url_for('payroll.index', month=month, year=year))

    if paid_days <= 0 or paid_days > emp_data['effective_days']:
        flash(f'Days must be between 0.5 and {emp_data["effective_days"]}.', 'error')
        return redirect(url_for('payroll.index', month=month, year=year))

    partial_amount = float(emp_data['daily_wage']) * paid_days
    pending_days = emp_data['effective_days'] - paid_days
    pending_amount = emp_data['net_salary'] - partial_amount

    existing = Payroll.query.filter_by(employee_id=employee_id, month=month, year=year).first()
    if existing:
        existing.present_days = paid_days
        existing.net_salary = partial_amount
        existing.gross_salary = emp_data['gross_salary']
        existing.advance_deduction = emp_data['advance_deduction']
        existing.is_paid = False
        existing.paid_at = datetime.now()
    else:
        record = Payroll(
            employee_id=employee_id, month=month, year=year,
            present_days=paid_days,
            daily_wage=emp_data['daily_wage'],
            gross_salary=emp_data['gross_salary'],
            advance_deduction=emp_data['advance_deduction'],
            net_salary=partial_amount,
            is_paid=False,
            paid_at=datetime.now()
        )
        db.session.add(record)

    db.session.commit()
    flash(f'Partial pay recorded: {paid_days} days (₹{partial_amount:,.0f}). Pending: {pending_days} days (₹{pending_amount:,.0f}).', 'success')
    return redirect(url_for('payroll.index', month=month, year=year))


@payroll_bp.route('/payroll/export')
@login_required
def export():
    today = date.today()
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))
    month_name = MONTH_NAMES[month - 1]

    payroll_data = calculate_payroll(year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Payroll {month_name} {year}'

    header_fill = PatternFill('solid', fgColor='1e2333')
    header_font = Font(bold=True, color='F5A623', name='Calibri', size=11)
    accent_font = Font(color='3ECF8E', bold=True, name='Calibri')
    danger_font = Font(color='E55353', name='Calibri')
    thin = Border(bottom=Side(style='thin', color='2a2f45'))

    headers = ['#', 'Employee Name', 'Daily Wage (₹)', 'Present Days',
               'Half Days', 'Effective Days', 'Gross Salary (₹)',
               'Advance Deduction (₹)', 'Net Salary (₹)', 'Paid Days', 'Pending Days', 'Status']

    col_widths = [5, 22, 15, 13, 10, 14, 17, 20, 15, 12, 13, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    for i, p in enumerate(payroll_data, 1):
        row = i + 1
        if p['is_paid']:
            status = 'Paid'
        elif p['is_partial']:
            status = 'Partial'
        else:
            status = 'Pending'

        values = [i, p['name'], p['daily_wage'], p['present_days'],
                  p['half_days'], p['effective_days'],
                  round(p['gross_salary'], 2), round(p['advance_deduction'], 2),
                  round(p['net_salary'], 2),
                  p['partial_days'] if p['is_partial'] else (p['effective_days'] if p['is_paid'] else 0),
                  p['pending_days'] if p['is_partial'] else 0,
                  status]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(vertical='center')
            if col == 9:
                cell.font = accent_font
            if col == 8 and p['advance_deduction'] > 0:
                cell.font = danger_font

    total_row = len(payroll_data) + 2
    ws.cell(row=total_row, column=2, value='TOTAL').font = Font(bold=True, name='Calibri')
    ws.cell(row=total_row, column=7, value=round(sum(p['gross_salary'] for p in payroll_data), 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=round(sum(p['advance_deduction'] for p in payroll_data), 2)).font = Font(bold=True, color='E55353')
    ws.cell(row=total_row, column=9, value=round(sum(p['net_salary'] for p in payroll_data), 2)).font = Font(bold=True, color='3ECF8E')

    ws2 = wb.create_sheet(title=f'Attendance {month_name} {year}')
    att_headers = ['Employee Name', 'Date', 'Status']
    for col, h in enumerate(att_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 12

    att_records = db.session.query(Attendance, Employee.name).join(Employee).filter(
        func.extract('year', Attendance.date) == year,
        func.extract('month', Attendance.date) == month
    ).order_by(Employee.name, Attendance.date).all()

    for i, (att, emp_name) in enumerate(att_records, 2):
        ws2.cell(row=i, column=1, value=emp_name)
        ws2.cell(row=i, column=2, value=att.date.strftime('%d-%m-%Y'))
        status_cell = ws2.cell(row=i, column=3, value=att.status.replace('_', ' ').title())
        if att.status == 'present':
            status_cell.font = Font(color='3ECF8E')
        elif att.status == 'absent':
            status_cell.font = Font(color='E55353')
        else:
            status_cell.font = Font(color='F5A623')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'payroll_{month_name}_{year}.xlsx'
    )
